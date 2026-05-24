#!/usr/bin/env python3
"""
Reverse CSV exports back into an XLSX workbook by using backlog/backlog.xlsx
as the formatting/template source.

Usage from the repository root:
    python reverse_convert.py

Default inputs:
    backlog/backlog.xlsx
    backlog/backlog.csv                 # convenience CSV for a one-sheet workbook
    backlog/csv/<safe_sheet_name>.csv   # per-sheet CSV exports

Default output:
    backlog/backlog_restored.xlsx

What this can restore:
    - Existing workbook formatting from backlog/backlog.xlsx
    - Existing sheets, column widths, row heights, merged cells, colors, borders,
      number formats, filters, frozen panes, and other workbook styling
    - Existing formulas by default, because they still exist in the template XLSX

What CSV cannot restore by itself:
    - Formatting that is not already present in the template workbook
    - Deleted sheets
    - Merged-cell structure if the template workbook is missing
    - Exact blank-row positions if the CSV was manually rearranged
    - Formula definitions if --overwrite-formulas is used

Notes:
    - Requires openpyxl:
        pip install openpyxl
    - The original backlog/backlog.xlsx is never overwritten unless --in-place is passed.
    - Rows are mapped onto the template's non-empty rows by default. This matches
      convert.py, which skipped fully empty rows when exporting to CSV.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell, MergedCell
except ImportError:
    print(
        "Missing dependency: openpyxl\n\n"
        "Install it with:\n"
        "    pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
DEFAULT_TEMPLATE_XLSX = REPO_ROOT / "backlog" / "backlog.xlsx"
DEFAULT_SINGLE_SHEET_CSV = REPO_ROOT / "backlog" / "backlog.csv"
DEFAULT_CSV_DIR = REPO_ROOT / "backlog" / "csv"
DEFAULT_OUTPUT_XLSX = REPO_ROOT / "backlog" / "backlog_restored.xlsx"


def safe_filename(name: str) -> str:
    """
    Match convert.py's Excel-sheet-name-to-CSV-file-name conversion.
    """
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip())
    cleaned = cleaned.strip("._-")
    return cleaned or "sheet"


def is_formula_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_empty_value(value: object) -> bool:
    return value is None or value == ""


def is_writable_cell(cell: Cell | MergedCell) -> bool:
    """
    openpyxl represents non-anchor cells in merged ranges as read-only MergedCell
    objects. The top-left anchor cell is a normal Cell and remains writable.
    """
    return not isinstance(cell, MergedCell)


def row_has_template_value(worksheet, row_number: int) -> bool:
    """
    convert.py skipped fully empty rows. To reverse that more faithfully, map CSV
    row N back to the Nth non-empty row in the template sheet.
    """
    for cell in worksheet[row_number]:
        if not is_writable_cell(cell):
            continue
        if not is_empty_value(cell.value):
            return True
    return False


def template_data_rows(worksheet) -> list[int]:
    rows = [
        row_number
        for row_number in range(1, worksheet.max_row + 1)
        if row_has_template_value(worksheet, row_number)
    ]
    return rows or list(range(1, worksheet.max_row + 1))


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    with csv_path.open("r", newline="", encoding="utf-8-sig") as csv_file:
        return [row for row in csv.reader(csv_file)]


def coerce_to_template_type(raw_value: str, template_value: object) -> object:
    """
    CSV stores everything as text. When the template cell already has a type,
    reuse that type where conversion is safe. If conversion fails, keep text.

    Empty CSV fields become blank cells.
    """
    if raw_value == "":
        return None

    if template_value is None or is_formula_value(template_value):
        return raw_value

    if isinstance(template_value, bool):
        normalized = raw_value.strip().lower()
        if normalized in {"true", "1", "yes", "y"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        return raw_value

    if isinstance(template_value, int) and not isinstance(template_value, bool):
        try:
            return int(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, float):
        try:
            return float(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, datetime):
        try:
            return datetime.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, date) and not isinstance(template_value, datetime):
        try:
            return date.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, time):
        try:
            return time.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    return raw_value


def clear_row_values(
    worksheet,
    row_number: int,
    *,
    max_column: int,
    preserve_formulas: bool,
) -> None:
    """
    Clear values from an existing template row before writing CSV values into it.
    Styles and formatting are preserved.
    """
    for column_number in range(1, max_column + 1):
        cell = worksheet.cell(row=row_number, column=column_number)

        if not is_writable_cell(cell):
            continue

        if preserve_formulas and is_formula_value(cell.value):
            continue

        cell.value = None


def write_csv_rows_to_sheet(
    worksheet,
    csv_rows: Sequence[Sequence[str]],
    *,
    preserve_formulas: bool,
    use_template_row_map: bool,
) -> int:
    """
    Write CSV rows into one worksheet while preserving worksheet formatting.
    Returns the number of CSV rows written.
    """
    mapped_rows = template_data_rows(worksheet) if use_template_row_map else []
    csv_width = max((len(row) for row in csv_rows), default=0)
    clear_width = max(worksheet.max_column, csv_width)

    for csv_row_index, csv_row in enumerate(csv_rows, start=1):
        if use_template_row_map and csv_row_index <= len(mapped_rows):
            worksheet_row = mapped_rows[csv_row_index - 1]
        else:
            if use_template_row_map and mapped_rows:
                worksheet_row = mapped_rows[-1] + (csv_row_index - len(mapped_rows))
            else:
                worksheet_row = csv_row_index

        clear_row_values(
            worksheet,
            worksheet_row,
            max_column=clear_width,
            preserve_formulas=preserve_formulas,
        )

        for column_number, raw_value in enumerate(csv_row, start=1):
            cell = worksheet.cell(row=worksheet_row, column=column_number)

            if not is_writable_cell(cell):
                continue

            if preserve_formulas and is_formula_value(cell.value):
                continue

            cell.value = coerce_to_template_type(raw_value, cell.value)

    return len(csv_rows)


def find_duplicate_safe_names(sheet_names: Iterable[str]) -> dict[str, list[str]]:
    by_safe_name: dict[str, list[str]] = {}

    for sheet_name in sheet_names:
        by_safe_name.setdefault(safe_filename(sheet_name), []).append(sheet_name)

    return {
        csv_stem: names
        for csv_stem, names in by_safe_name.items()
        if len(names) > 1
    }


def csv_for_sheet(
    *,
    sheet_title: str,
    visible_sheet_count: int,
    single_sheet_csv: Path,
    csv_dir: Path,
) -> Path | None:
    """
    Match convert.py's outputs:
      - backlog/backlog.csv for a one-visible-sheet workbook
      - backlog/csv/<safe_sheet_name>.csv for each visible sheet
    """
    if visible_sheet_count == 1 and single_sheet_csv.exists():
        return single_sheet_csv

    per_sheet_csv = csv_dir / f"{safe_filename(sheet_title)}.csv"
    if per_sheet_csv.exists():
        return per_sheet_csv

    return None


def force_recalculation_on_open(workbook) -> None:
    """
    Ask spreadsheet applications to recalculate formulas when the output workbook
    is opened. Different openpyxl versions expose calculation properties with
    slightly different attributes, so this is intentionally best-effort.
    """
    calculation = getattr(workbook, "calculation", None)

    if calculation is None:
        return

    for attribute, value in {
        "fullCalcOnLoad": True,
        "forceFullCalc": True,
        "calcMode": "auto",
    }.items():
        try:
            setattr(calculation, attribute, value)
        except Exception:
            pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Rebuild an XLSX file from CSV exports by writing CSV values into "
            "the existing backlog/backlog.xlsx template."
        )
    )

    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE_XLSX,
        help=f"Template workbook to preserve formatting from. Default: {DEFAULT_TEMPLATE_XLSX}",
    )
    parser.add_argument(
        "--single-sheet-csv",
        type=Path,
        default=DEFAULT_SINGLE_SHEET_CSV,
        help=f"CSV used when the template has one visible sheet. Default: {DEFAULT_SINGLE_SHEET_CSV}",
    )
    parser.add_argument(
        "--csv-dir",
        type=Path,
        default=DEFAULT_CSV_DIR,
        help=f"Directory containing per-sheet CSV files. Default: {DEFAULT_CSV_DIR}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_XLSX,
        help=f"Output workbook path. Default: {DEFAULT_OUTPUT_XLSX}",
    )
    parser.add_argument(
        "--in-place",
        action="store_true",
        help="Overwrite the template workbook instead of writing --output.",
    )
    parser.add_argument(
        "--overwrite-formulas",
        action="store_true",
        help=(
            "Replace template formulas with CSV values. By default, formulas in "
            "the template workbook are preserved."
        ),
    )
    parser.add_argument(
        "--sequential-rows",
        action="store_true",
        help=(
            "Write CSV rows to rows 1..N instead of mapping them to non-empty "
            "template rows. The default better matches convert.py's skipped-empty-row export."
        ),
    )

    return parser.parse_args()


def main() -> int:
    args = parse_args()

    template_path = args.template.resolve()
    output_path = template_path if args.in_place else args.output.resolve()
    single_sheet_csv = args.single_sheet_csv.resolve()
    csv_dir = args.csv_dir.resolve()

    if not template_path.exists():
        print(f"Could not find template workbook: {template_path}", file=sys.stderr)
        return 1

    if output_path == template_path and not args.in_place:
        print(
            "Refusing to overwrite the template workbook without --in-place.",
            file=sys.stderr,
        )
        return 1

    workbook = load_workbook(template_path)
    visible_sheets = [
        sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"
    ]

    if not visible_sheets:
        print(f"No visible sheets found in: {template_path}", file=sys.stderr)
        return 1

    duplicates = find_duplicate_safe_names(sheet.title for sheet in visible_sheets)
    if duplicates:
        print(
            "Cannot safely map CSV files because multiple visible sheets produce "
            "the same safe CSV filename:",
            file=sys.stderr,
        )
        for csv_stem, names in duplicates.items():
            print(f"  {csv_stem}.csv: {', '.join(names)}", file=sys.stderr)
        return 1

    preserve_formulas = not args.overwrite_formulas
    use_template_row_map = not args.sequential_rows

    print(f"Template: {template_path}")
    print(f"Output:   {output_path}")
    print(f"Mode:     {'preserve formulas' if preserve_formulas else 'overwrite formulas'}")
    print()

    sheets_written = 0

    for worksheet in visible_sheets:
        csv_path = csv_for_sheet(
            sheet_title=worksheet.title,
            visible_sheet_count=len(visible_sheets),
            single_sheet_csv=single_sheet_csv,
            csv_dir=csv_dir,
        )

        if csv_path is None:
            print(f"Skipping sheet without matching CSV: {worksheet.title}")
            continue

        csv_rows = read_csv_rows(csv_path)
        row_count = write_csv_rows_to_sheet(
            worksheet,
            csv_rows,
            preserve_formulas=preserve_formulas,
            use_template_row_map=use_template_row_map,
        )

        print(f"Wrote {row_count} CSV rows into sheet '{worksheet.title}' from: {csv_path}")
        sheets_written += 1

    if sheets_written == 0:
        print(
            "No sheets were written. Expected either backlog/backlog.csv for a "
            "single visible sheet or backlog/csv/<safe_sheet_name>.csv files.",
            file=sys.stderr,
        )
        return 1

    force_recalculation_on_open(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print()
    print(f"Done. Wrote: {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
