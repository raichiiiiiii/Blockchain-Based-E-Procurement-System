#!/usr/bin/env python3
"""
Convert backlog/backlog.xlsx to CSV without passing any filename.

Usage from the repository root:
    python convert.py

Default input:
    backlog/backlog.xlsx

Output:
    - If workbook has one visible sheet:
        backlog/backlog.csv

    - For every visible sheet:
        backlog/csv/<safe_sheet_name>.csv

Notes:
    - Requires openpyxl:
        pip install openpyxl
    - Uses cached formula results where available.
    - CSV cannot preserve Excel formatting, merged cells, formulas, colors, or multiple sheets
      in a single file, so this script writes one CSV per visible sheet.
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Iterable, Any

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Missing dependency: openpyxl\n\n"
        "Install it with:\n"
        "    pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
INPUT_XLSX = REPO_ROOT / "backlog" / "backlog.xlsx"
OUTPUT_DIR = REPO_ROOT / "backlog" / "csv"
SINGLE_SHEET_OUTPUT = REPO_ROOT / "backlog" / "backlog.csv"


def safe_filename(name: str) -> str:
    """
    Convert an Excel sheet name into a filesystem-safe CSV filename.
    """
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name.strip())
    cleaned = cleaned.strip("._-")
    return cleaned or "sheet"


def cell_to_csv_value(value: Any) -> str:
    """
    Convert Excel cell values into stable CSV text.

    Dates/numbers/booleans are handled by csv.writer after string conversion.
    None becomes an empty field.
    """
    if value is None:
        return ""
    return str(value)


def trim_trailing_empty_cells(row: Iterable[Any]) -> list[str]:
    """
    Remove empty cells at the end of a row so CSV output is cleaner.
    Internal empty cells are preserved.
    """
    values = [cell_to_csv_value(value) for value in row]

    while values and values[-1] == "":
        values.pop()

    return values


def write_sheet_to_csv(worksheet, output_path: Path) -> int:
    """
    Write one worksheet to one CSV file.
    Returns the number of written rows.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows_written = 0

    with output_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)

        for row in worksheet.iter_rows(values_only=True):
            cleaned_row = trim_trailing_empty_cells(row)

            # Skip fully empty rows.
            if not cleaned_row:
                continue

            writer.writerow(cleaned_row)
            rows_written += 1

    return rows_written


def main() -> int:
    if not INPUT_XLSX.exists():
        print(f"Could not find expected workbook: {INPUT_XLSX}", file=sys.stderr)
        print("Run this script from the repository root, or place it in the repository root.", file=sys.stderr)
        return 1

    workbook = load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]

    if not visible_sheets:
        print(f"No visible sheets found in: {INPUT_XLSX}", file=sys.stderr)
        return 1

    print(f"Reading: {INPUT_XLSX}")

    generated_files: list[Path] = []

    for sheet in visible_sheets:
        csv_name = f"{safe_filename(sheet.title)}.csv"
        csv_path = OUTPUT_DIR / csv_name
        row_count = write_sheet_to_csv(sheet, csv_path)
        generated_files.append(csv_path)
        print(f"Wrote {row_count} rows: {csv_path}")

    if len(visible_sheets) == 1:
        row_count = write_sheet_to_csv(visible_sheets[0], SINGLE_SHEET_OUTPUT)
        generated_files.append(SINGLE_SHEET_OUTPUT)
        print(f"Wrote single-sheet convenience copy: {SINGLE_SHEET_OUTPUT}")

    print("\nDone.")
    print("Generated:")
    for path in generated_files:
        print(f"  - {path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
