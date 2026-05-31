#!/usr/bin/env python3
"""
Create an XLSX copy from backlog/backlog.csv while preserving the formatting from
backlog/backlog.xlsx.

Run from the repository root:
    python csv_to_xlsx.py

Defaults:
    CSV input:      backlog/backlog.csv
    XLSX template:  backlog/backlog.xlsx
    XLSX output:    backlog/backlog_from_csv.xlsx

The template workbook is never overwritten unless --output points to it explicitly
and --allow-template-overwrite is passed.
"""

from __future__ import annotations

import argparse
import csv
import sys
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Sequence

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError:
    print(
        "Missing dependency: openpyxl\n\n"
        "Install it with:\n"
        "    pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
DEFAULT_CSV = REPO_ROOT / "backlog" / "backlog.csv"
DEFAULT_TEMPLATE = REPO_ROOT / "backlog" / "backlog.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "backlog" / "backlog_from_csv.xlsx"


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    """Read CSV rows, accepting UTF-8 with or without a BOM."""
    with csv_path.open("r", newline="", encoding="utf-8-sig") as csv_file:
        return [row for row in csv.reader(csv_file)]


def is_empty(value: Any) -> bool:
    return value is None or value == ""


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_writable_cell(cell: Cell | MergedCell) -> bool:
    """Merged-range cells other than the top-left anchor are read-only."""
    return not isinstance(cell, MergedCell)


def visible_worksheets(workbook) -> list[Any]:
    return [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]


def choose_target_sheet(workbook, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Template workbook does not contain sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        if worksheet.sheet_state != "visible":
            raise ValueError(f"Target sheet is not visible: {sheet_name}")
        return worksheet

    visible = visible_worksheets(workbook)
    if not visible:
        raise ValueError("Template workbook has no visible sheets.")
    return visible[0]


def snapshot_template_values(worksheet, max_row: int, max_column: int) -> dict[tuple[int, int], Any]:
    """Capture original values before clearing so type coercion can use them."""
    return {
        (row_number, column_number): worksheet.cell(row_number, column_number).value
        for row_number in range(1, max_row + 1)
        for column_number in range(1, max_column + 1)
    }


def copy_cell_format(source: Cell, target: Cell) -> None:
    """Copy formatting from one cell to another without copying the value."""
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format

    if source.style:
        target.style = source.style


def copy_row_format(worksheet, source_row_number: int, target_row_number: int) -> None:
    """
    Copy row-level formatting.

    RowDimension.customHeight is a read-only computed property in openpyxl, so it
    must not be assigned directly. Setting height is enough for openpyxl to mark a
    custom row height when one exists.
    """
    source_dim = worksheet.row_dimensions[source_row_number]
    target_dim = worksheet.row_dimensions[target_row_number]

    target_dim.height = source_dim.height
    target_dim.hidden = source_dim.hidden
    target_dim.outlineLevel = source_dim.outlineLevel
    target_dim.collapsed = source_dim.collapsed


def source_row_for(target_row: int, template_max_row: int) -> int:
    if template_max_row <= 0:
        return 1
    return min(target_row, template_max_row)


def source_column_for(target_column: int, template_max_column: int) -> int:
    if template_max_column <= 0:
        return 1
    return min(target_column, template_max_column)


def coerce_to_template_type(raw_value: str, template_value: Any) -> Any:
    """
    CSV values are text. Reuse the template cell's type where conversion is safe.
    Empty CSV fields become blank cells.
    """
    if raw_value == "":
        return None

    if template_value is None or is_formula(template_value):
        return raw_value

    if isinstance(template_value, bool):
        normalized = raw_value.strip().lower()
        if normalized in {"true", "1", "yes", "y"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        return raw_value

    if isinstance(template_value, int) and not isinstance(template_value, bool):
        try:
            return int(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, float):
        try:
            return float(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, datetime):
        try:
            return datetime.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, date) and not isinstance(template_value, datetime):
        try:
            return date.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    if isinstance(template_value, time):
        try:
            return time.fromisoformat(raw_value)
        except ValueError:
            return raw_value

    return raw_value


def clear_existing_values(
    worksheet,
    *,
    rows_to_clear: int,
    columns_to_clear: int,
    preserve_formulas: bool,
) -> None:
    """Clear cell values while leaving formatting in place."""
    for row_number in range(1, rows_to_clear + 1):
        for column_number in range(1, columns_to_clear + 1):
            cell = worksheet.cell(row=row_number, column=column_number)
            if not is_writable_cell(cell):
                continue
            if preserve_formulas and is_formula(cell.value):
                continue
            cell.value = None


def apply_template_format_for_target_cell(
    worksheet,
    *,
    target_row: int,
    target_column: int,
    template_max_row: int,
    template_max_column: int,
) -> None:
    """Apply the nearest template cell's style to a target cell."""
    source_row = source_row_for(target_row, template_max_row)
    source_column = source_column_for(target_column, template_max_column)

    if source_row == target_row and source_column == target_column:
        return

    source_cell = worksheet.cell(row=source_row, column=source_column)
    target_cell = worksheet.cell(row=target_row, column=target_column)

    if is_writable_cell(target_cell) and isinstance(source_cell, Cell):
        copy_cell_format(source_cell, target_cell)


def write_rows_to_template_sheet(
    worksheet,
    csv_rows: Sequence[Sequence[str]],
    *,
    preserve_formulas: bool,
) -> None:
    template_max_row = max(worksheet.max_row, 1)
    template_max_column = max(worksheet.max_column, 1)
    csv_row_count = len(csv_rows)
    csv_column_count = max((len(row) for row in csv_rows), default=0)
    max_column = max(template_max_column, csv_column_count)
    max_row = max(template_max_row, csv_row_count)

    template_values = snapshot_template_values(worksheet, template_max_row, template_max_column)

    clear_existing_values(
        worksheet,
        rows_to_clear=max_row,
        columns_to_clear=max_column,
        preserve_formulas=preserve_formulas,
    )

    for row_number in range(1, csv_row_count + 1):
        source_row = source_row_for(row_number, template_max_row)
        if row_number > template_max_row:
            copy_row_format(worksheet, source_row, row_number)

        for column_number in range(1, max_column + 1):
            apply_template_format_for_target_cell(
                worksheet,
                target_row=row_number,
                target_column=column_number,
                template_max_row=template_max_row,
                template_max_column=template_max_column,
            )

        csv_row = csv_rows[row_number - 1]
        for column_number, raw_value in enumerate(csv_row, start=1):
            cell = worksheet.cell(row=row_number, column=column_number)
            if not is_writable_cell(cell):
                continue
            if preserve_formulas and is_formula(cell.value):
                continue

            source_row = source_row_for(row_number, template_max_row)
            source_column = source_column_for(column_number, template_max_column)
            template_value = template_values.get((source_row, source_column))
            cell.value = coerce_to_template_type(raw_value, template_value)


def resize_tables_and_filters(worksheet, row_count: int, column_count: int) -> None:
    """
    Keep Excel table and autofilter ranges aligned with the CSV data size.
    This assumes the CSV writes a rectangular dataset beginning at A1.
    """
    if row_count < 1 or column_count < 1:
        return

    full_ref = f"A1:{get_column_letter(column_count)}{row_count}"

    if worksheet.auto_filter and worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = full_ref

    for table in worksheet.tables.values():
        min_col, min_row, _max_col, _max_row = range_boundaries(table.ref)
        new_max_col = min_col + column_count - 1
        new_max_row = min_row + row_count - 1
        table.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(new_max_col)}{new_max_row}"
        )
        if table.autoFilter:
            table.autoFilter.ref = table.ref


def force_recalculation_on_open(workbook) -> None:
    """Ask Excel-compatible apps to recalculate formulas when the workbook opens."""
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return

    for attribute, value in {
        "fullCalcOnLoad": True,
        "forceFullCalc": True,
        "calcMode": "auto",
    }.items():
        try:
            setattr(calculation, attribute, value)
        except Exception:
            pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert backlog/backlog.csv into an XLSX copy using "
            "backlog/backlog.xlsx as the formatting template."
        )
    )
    parser.add_argument(
        "--csv",
        dest="csv_path",
        type=Path,
        default=DEFAULT_CSV,
        help=f"CSV file to convert. Default: {DEFAULT_CSV}",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=DEFAULT_TEMPLATE,
        help=f"XLSX template to copy formatting from. Default: {DEFAULT_TEMPLATE}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"XLSX output file to create. Default: {DEFAULT_OUTPUT}",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Target sheet name. Defaults to the first visible sheet in the template.",
    )
    parser.add_argument(
        "--preserve-formulas",
        action="store_true",
        help="Keep existing template formulas instead of replacing them with CSV values.",
    )
    parser.add_argument(
        "--allow-template-overwrite",
        action="store_true",
        help="Allow --output to be the same file as --template.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    csv_path = args.csv_path.resolve()
    template_path = args.template.resolve()
    output_path = args.output.resolve()

    if not csv_path.exists():
        print(f"Could not find CSV input: {csv_path}", file=sys.stderr)
        return 1

    if not template_path.exists():
        print(f"Could not find XLSX template: {template_path}", file=sys.stderr)
        return 1

    if output_path == template_path and not args.allow_template_overwrite:
        print(
            "Refusing to overwrite the template workbook. Choose a different --output "
            "or pass --allow-template-overwrite.",
            file=sys.stderr,
        )
        return 1

    csv_rows = read_csv_rows(csv_path)
    if not csv_rows:
        print(f"CSV input has no rows: {csv_path}", file=sys.stderr)
        return 1

    workbook = load_workbook(template_path)

    try:
        worksheet = choose_target_sheet(workbook, args.sheet)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    write_rows_to_template_sheet(
        worksheet,
        csv_rows,
        preserve_formulas=args.preserve_formulas,
    )
    resize_tables_and_filters(
        worksheet,
        row_count=len(csv_rows),
        column_count=max((len(row) for row in csv_rows), default=0),
    )
    force_recalculation_on_open(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"CSV input:     {csv_path}")
    print(f"XLSX template: {template_path}")
    print(f"Target sheet:  {worksheet.title}")
    print(f"Rows written:  {len(csv_rows)}")
    print(f"Output XLSX:   {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
