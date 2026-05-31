#!/usr/bin/env python3
"""
Overwrite backlog/backlog.csv from backlog/backlog.xlsx.

Run from the repository root:
    python xlsx_to_csv.py

Defaults:
    XLSX input:  backlog/backlog.xlsx
    CSV output:  backlog/backlog.csv

Notes:
    - Requires openpyxl:
        pip install openpyxl
    - Uses cached formula results where available.
    - CSV cannot preserve workbook formatting, colors, formulas, merged cells, or
      multiple sheets. This script writes the selected worksheet's values only.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Missing dependency: openpyxl\n\n"
        "Install it with:\n"
        "    pip install openpyxl\n",
        file=sys.stderr,
    )
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = REPO_ROOT / "backlog" / "backlog.xlsx"
DEFAULT_CSV = REPO_ROOT / "backlog" / "backlog.csv"


def visible_worksheets(workbook) -> list[Any]:
    return [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]


def choose_source_sheet(workbook, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook does not contain sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        if worksheet.sheet_state != "visible":
            raise ValueError(f"Source sheet is not visible: {sheet_name}")
        return worksheet

    visible = visible_worksheets(workbook)
    if not visible:
        raise ValueError("Workbook has no visible sheets.")
    return visible[0]


def cell_to_csv_value(value: Any) -> str:
    """Convert worksheet cell values into stable CSV text."""
    if value is None:
        return ""
    return str(value)


def trim_trailing_empty_cells(row: Iterable[Any]) -> list[str]:
    """Remove trailing empty cells while preserving internal empty cells."""
    values = [cell_to_csv_value(value) for value in row]

    while values and values[-1] == "":
        values.pop()

    return values


def worksheet_rows_for_csv(worksheet) -> list[list[str]]:
    rows: list[list[str]] = []

    for row in worksheet.iter_rows(values_only=True):
        cleaned_row = trim_trailing_empty_cells(row)
        if not cleaned_row:
            continue
        rows.append(cleaned_row)

    return rows


def write_csv_atomic(rows: list[list[str]], output_path: Path) -> None:
    """Write the CSV through a temporary file, then replace the output path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fd, temp_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}.",
        suffix=output_path.suffix,
        dir=output_path.parent,
        text=True,
    )
    os.close(fd)
    temp_path = Path(temp_name)

    try:
        with temp_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(rows)
        os.replace(temp_path, output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert backlog/backlog.xlsx into backlog/backlog.csv."
    )
    parser.add_argument(
        "--xlsx",
        dest="xlsx_path",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"XLSX workbook to read. Default: {DEFAULT_XLSX}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_CSV,
        help=f"CSV file to write. Default: {DEFAULT_CSV}",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Source sheet name. Defaults to the first visible sheet.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    xlsx_path = args.xlsx_path.resolve()
    output_path = args.output.resolve()

    if not xlsx_path.exists():
        print(f"Could not find XLSX input: {xlsx_path}", file=sys.stderr)
        return 1

    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)

    try:
        worksheet = choose_source_sheet(workbook, args.sheet)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    rows = worksheet_rows_for_csv(worksheet)
    if not rows:
        print(f"Source sheet has no non-empty rows: {worksheet.title}", file=sys.stderr)
        return 1

    write_csv_atomic(rows, output_path)

    print(f"XLSX input:   {xlsx_path}")
    print(f"Source sheet: {worksheet.title}")
    print(f"Rows written: {len(rows)}")
    print(f"CSV output:   {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
